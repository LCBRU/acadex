from acadex.services import add_or_update_academic
import csv
import io
from flask import Blueprint, request, render_template, redirect, url_for, make_response, send_file
from flask_security import login_required
from lbrc_flask.database import db
from lbrc_flask.forms import SearchForm
from openpyxl.styles import Font
from acadex.model import Academic, Publication
from scholarly import scholarly
from itertools import islice
from datetime import datetime
from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from flask_weasyprint import HTML, render_pdf
from .forms import PublicationSearchForm


blueprint = Blueprint("ui", __name__, template_folder="templates")

# Login required for all views
@blueprint.before_request
@login_required
def before_request():
    pass


@blueprint.record
def record(state):
    if db is None:
        raise Exception(
            "This blueprint expects you to provide " "database access through database"
        )

@blueprint.route("/")
def index():
    search_form = SearchForm(formdata=request.args)

    q = Academic.query

    if search_form.search.data:
        q = q.filter(Academic.name.like("%{}%".format(search_form.search.data)))

    q = q.order_by(Academic.name.asc())

    academics = q.paginate(
            page=search_form.page.data,
            per_page=5,
            error_out=False,
        )

    return render_template("ui/index.html", academics=academics, search_form=search_form)


@blueprint.route("/publications")
def publications():
    search_form = PublicationSearchForm(formdata=request.args)

    q = Publication.query

    if search_form.search.data:
        q = q.filter(Publication.title.like("%{}%".format(search_form.search.data)))

    if search_form.academic_id.data:
        q = q.join(Publication.academics)
        q = q.filter(Academic.id == search_form.academic_id.data)

    q = q.order_by(Publication.published_date.desc())

    publications = q.paginate(
            page=search_form.page.data,
            per_page=5,
            error_out=False,
        )

    return render_template("ui/publications.html", publications=publications, search_form=search_form)


@blueprint.route("/add_search")
def add_search():
    search_form = SearchForm(formdata=request.args)

    academics = []

    if search_form.search.data:
        for a in islice(scholarly.search_author(search_form.search.data), 0, 10):
            academics.append(a)

    return render_template("ui/add.html", academics=academics, search_form=search_form)


@blueprint.route("/add_or_update/<string:google_scholar_id>/", methods=['GET', 'POST'])
def add_or_update(google_scholar_id):
    print('Walk')

    add_or_update_academic(google_scholar_id)

    return redirect(url_for('ui.index'))


@blueprint.route('/download/csv')
def download_csv():

    COL_NAME = 'Name'
    COL_AFFILIATION = 'Affiliation'
    COL_CITATIONS = 'Citations'
    COL_H_INDEX = 'H-Index'
    COL_I10_INDEX = 'I10-Index'

    fieldnames = [
        COL_NAME,
        COL_AFFILIATION,
        COL_CITATIONS,
        COL_H_INDEX,
        COL_I10_INDEX,
    ]

    si = io.StringIO()

    output = csv.DictWriter(
        si,
        fieldnames=fieldnames,
        quoting=csv.QUOTE_NONNUMERIC
    )

    output.writeheader()
    
    for a in Academic.query.order_by(Academic.name).all():
        output.writerow({
            COL_NAME: a.name,
            COL_AFFILIATION: a.affiliation,
            COL_CITATIONS: a.cited_by,
            COL_H_INDEX: a.h_index,
            COL_I10_INDEX: a.i10_index,
    })

    resp = make_response(si.getvalue().encode('utf-8'))
    resp.headers["Content-Disposition"] = "attachment; filename=acadex_{}.csv".format(datetime.utcnow().strftime("%c"))
    resp.headers["Content-type"] = "text/csv; charset=utf-8"
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = 0
    return resp


@blueprint.route('/download/excel')
def download_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Academics"
    bold = Font(bold=True)

    ws['A1'].style = 'Headline 1'
    ws['B1'].style = 'Headline 1'
    ws['C1'].style = 'Headline 1'
    ws['D1'].style = 'Headline 1'
    ws['E1'].style = 'Headline 1'

    ws.cell(column=1, row=1, value='Name')
    ws.cell(column=2, row=1, value='Affiliation')
    ws.cell(column=3, row=1, value='Citations')
    ws.cell(column=4, row=1, value='H-Index')
    ws.cell(column=5, row=1, value='I10-index')

    for row, a in enumerate(Academic.query.order_by(Academic.name).all(), start=2):
        ws['A{}'.format(row)].font = bold
        ws.cell(column=1, row=row, value=a.name)
        ws.cell(column=2, row=row, value=a.affiliation)
        ws.cell(column=3, row=row, value=a.cited_by)
        ws.cell(column=4, row=row, value=a.h_index)
        ws.cell(column=5, row=row, value=a.i10_index)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.flush()
        return send_file(
            tmp.name,
            as_attachment=True,
            attachment_filename='acadex_{}.xlsx'.format(datetime.utcnow().strftime("%Y%m%d_%H%M%S")),
            cache_timeout=0,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )


@blueprint.route('/download/pdf')
def download_pdf():
    academics = Academic.query.order_by(Academic.name).all()

    html = render_template('ui/pdf.html', academics=academics)
    return render_pdf(HTML(string=html), download_filename='acadex_{}.pdf'.format(datetime.utcnow().strftime("%Y%m%d_%H%M%S")))
